import gc
import logging
import pathlib
import matlab.engine
import matlab
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils.cell import coordinate_to_tuple
import io
from threading import Lock


def starting_table_number(table: Table):
    return coordinate_to_tuple(table.ref.split(":")[0])


def parse_tables(tables_to_consider: int, workbook: Workbook, *sheets: str):
    """
    Parse the tables in the workbook.

    :param tables_to_consider: The Top n tables to consider.
    :param workbook: The workbook to parse.
    :param sheets: The sheets to parse.
    :return: The parsed tables.
    """

    for SHEET in sheets:
        _sheet = workbook[SHEET]
        _tables = sorted(_sheet.tables.values(), key=lambda _: starting_table_number(_))

        logging.info("Parsing the sheet %s", SHEET)

        if len(_tables) < tables_to_consider:
            raise ValueError(
                f"Uploaded Workbook contains Sheet: `{SHEET}` which has less than {tables_to_consider} tables."
                f"It is important to arrange the tables in a sheet. Please know more about this in the docs."
            )

        for table in _tables[:tables_to_consider]:
            logging.info("For sheet %s, parsing the table: %s", SHEET, table.displayName)

            table: Table = table
            yield [
                [float(str(CELL.value).strip()) for CELL in row]
                for index, row in enumerate(_sheet[table.ref]) if index != 0
            ]

            logging.info("Table %s of Sheet %s was parsed successfully", SHEET, table.displayName)


class Engine:
    def __init__(self):
        super().__init__()
        self.matlab_engine = None
        self.future_response = None
        self.loading = Lock()
        self.processing = Lock()
        self.raw = []
        self._engine_pid = None

    def load(self):
        with self.loading:
            if self.future_response:
                return ...

            logging.info("Loading Engine...")
            self.future_response = matlab.engine.start_matlab(background=True)  # started async
        return ...

    @property
    def engine(self):
        if self.matlab_engine:
            return self.matlab_engine

        if not self.future_response:
            self.load()

        self.matlab_engine = self.future_response.result()
        self.matlab_engine.cd(str(pathlib.Path(__file__).parent / "Scripts"))
        self._engine_pid = self.matlab_engine.feature("GetPid")
        return self.matlab_engine

    def stop_engine(self):
        self.matlab_engine.quit() if self.matlab_engine else ...
        del self.matlab_engine
        gc.collect()

    def actual_parse(self, saved_excel: pathlib.Path, exams: int):
        workbook = load_workbook(saved_excel, read_only=False)
        logging.info("Workbook loaded")

        output = io.StringIO()
        error = io.StringIO()

        if len(workbook.sheetnames) < (exams + 1):
            raise ValueError(
                f"Uploaded Workbook has following sheets: {workbook.sheetnames}, "
                f"which is of length: {len(workbook.sheetnames)} it is required to be at least {exams + 1}. "
                f"Please refer the documentation in the Repo. or look for Help in Application"
            )

        sheets = workbook.sheetnames[-(exams + 1):]
        # cot, co_po, weightage, expected = (matlab.double(_) for _ in parse_tables(4, workbook, sheets[0]))
        gc.collect()

        self.engine.bridge(
            *(matlab.double(_) for _ in parse_tables(4, workbook, sheets[0])),
            *(matlab.double(_) for _ in parse_tables(2, workbook, *sheets[1:])),
            stdout=output,
            stderr=error,
            nargout=0
        )

        workbook.close()

        output.seek(0)
        error.seek(0)

        if saved_excel.exists():
            saved_excel.unlink()
            parent = saved_excel.parent
            if parent.exists() and not list(parent.iterdir()):
                parent.rmdir()

        return output.read(), error.read()

    def parse(self, *args):
        if self.processing.locked():
            return False, False

        with self.processing:
            return self.actual_parse(*args)

#
# if __name__ == "__main__":
#     engine = Engine()
#     engine.load()
#     a, b = engine.actual_parse("Downloads\\OBE IES.xlsx", 3)
#     print(a, b)
#     engine.stop_engine()
